

import os
from pathlib import Path
import pandas as pd 
import geopandas as gpd 
import openpyxl
from openpyxl.styles import (
    PatternFill,
    Border, 
    Side, 
    Font,
)


def addStyle(path:str, sheets:list, classes:list, colors:dict=None, sheet_type:str='matrix'):
    """
    addStyle This method identifies cells in the matrices that are unlikely and need verification (yellow) and incorrect
             transitions (red) and color codes them.

    Parameters
    ----------
    path : str
        path to excel workbook containing LCC matrices.
    sheets : list
        list of excel sheet names.
    classes : list
        list of unique LC classes
    colors : dict, optional, default is None
        dictionary of LC transitions and the color to assign.
    sheet_type : str, optional, default is matrix
        Sheet type. Options: matrix, difference, totals
    """     

    # excel
    worksheet = openpyxl.load_workbook(path)
    for i in range(len(sheets)):
        cur_sheet = worksheet[sheets[i]]

        if sheet_type =='matrix':
            for color in colors:
                # pattern to fill 
                fmtFillPattern = PatternFill(start_color=colors[color]['hex'],fill_type='solid')

                # iterate transitions
                xy = []
                for t in colors[color]['transitions']:
                    x, y = classes.index(t[0])+2, classes.index(t[1])+2
                    xy.append((x,y))

                # Apply fill
                for row in cur_sheet.iter_rows(cur_sheet.min_row,cur_sheet.max_row):
                    for cell in row:
                        if (cell.row, cell.col_idx) in xy:
                            if cell.value > 0:
                                cell.fill = fmtFillPattern
        
        # TODO add style to totals and differenced matrix

        # apply whole number format and borders around all cells
        thin = Side(border_style="thin", color="000000")
        for row in cur_sheet.iter_rows(cur_sheet.min_row,cur_sheet.max_row):
            for cell in row:  
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)  
                if cell.column > 1: # skip left index

                    if sheet_type == 'difference': # represents a percentage, negative numbers are parenthesis, 0 is a dash
                        cell.number_format = '_(* 0.00%_);_(* (0.00%);_(* "-"??_);_(@_)'
                    else:
                        # one decimal point, show commas, negative numbers are parenthesis, 0 is a dash
                        cell.number_format = '_(* #,##0.0_);_(* (#,##0.0);_(* "-"??_);_(@_)' # display as whole number without rounding true values
                    
                    # bold totals columns
                    if cur_sheet.cell(cell.row, 1).value in ['Decrease', 'Increase', 'Net Change']:
                        cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=True)

        # adjust column size
        max_length = 12
        if sheet_type == 'totals':
            max_length = 18
        for col in cur_sheet.columns:
            column = col[0].column_letter # Get the column name
            if sheet_type == 'totals' and col[0].value == 'LandCover':
                cur_sheet.column_dimensions[column].width = 35
            else:
                cur_sheet.column_dimensions[column].width = max_length

    # save formatting
    worksheet.save(path)
    worksheet.close()

def createMatrices(
                data_folder:str,
                qaqc_path:str,
                cf:str,
                years:list,
                version:str,
                lcc_lookup:pd.DataFrame,
                colors:dict,
                lc_abbrev:dict=None) -> pd.DataFrame:
    """
    createMatrices This method locates the LCC raster attribute tables, converts them to change matrices, and
                   calls the addStyle method to color-code unlikely transitions. The values are in acres and 
                   are not rounded to allow for identification of very small (including single-pixel) transitions.

    Parameters
    ----------
    data_folder : str
        path to folder containing cf folders.
    qaqc_path : str
        path to write all QA results for the county.
    cfs : str
        county_fips AKA cofips AKA cfs.
    years : list
        list of years mapped for county.
    version : str
        data version (2022ed or 2024ed).
    lcc_lookup : pd.DataFrame
        dataframe relating lcc values and the lc change names.
    colors : dict
        dictionary of LC transitions and the color to assign
    lc_abbrev : dict, Optional, default None
        dictionary to abbreviate LC class names in matrices.
    """
    # locate LC change paths
    input_folder = f"{data_folder}/{cf}/input"

    # dataframe of LC totals
    totals = []

    # read in change raster RATs and convert to change matrix
    matrices = []
    for i in range(len(years)-1):
        # read in RAT
        p = f"{input_folder}/{cf}_landcoverchange_{years[i]}_{years[i+1]}.tif.vat.dbf"
        if version == '2022ed':
            p = f"{input_folder}/{cf}_landcoverchange_{years[i]}{years[i+1]}_v2.tif.vat.dbf"
        if not os.path.isfile(p):
            raise Exception(f"{Path(p).name} does not exist.")

        df = (
            gpd.read_file(p)
            .filter(items=['Value', 'Count'])
            .set_index('Value')
        )

        # add LC names
        df = (
            df
            .merge(lcc_lookup, 
                   left_on='Value', 
                   right_index=True, 
                   how='left')
            .rename(columns = {
                'early' : str(years[i]),
                'late'  : str(years[i+1]),
            })
        )
        
        # convert count (square meters) to acres
        df.loc[:, "Acres"] = df['Count'] / 4046.86

        # calculate totals
        totals.append(
            totals_helper(df, str(years[i]), str(years[i+1]), version)
        )

        # if abbrevation dictionary - replace names with abbreviations
        if lc_abbrev is not None:
            df = (
                df
                .replace({str(years[i]):lc_abbrev})
                .replace({str(years[i+1]):lc_abbrev})
            )
            classes = list(lc_abbrev.values())
        else:
            # list of all unique classes
            classes = lcc_lookup['early'].unique().tolist()

        # drop cols
        df = df[[f"{years[i]}", f"{years[i+1]}", "Acres"]]

        # create pivot table
        matrix = (
            pd.pivot_table(df, values='Acres', index=f"{years[i]}", columns=f"{years[i+1]}")
            .round(4)
        )

        # validate all classes are present
        for c in classes:
            if c not in matrix.columns:
                matrix.loc[:, c] = 0
            if c not in matrix.index:
                matrix.loc[c, :] = 0

            # ensure no change is 0
            matrix.loc[c, c] = 0
        
        # change class order
        matrix = matrix.reindex(classes)
        matrix = matrix[classes]
        matrix = matrix.fillna(0)

        # totals
        matrix.loc[:, 'Decrease'] = matrix[classes].sum(axis=1)
        matrix.loc['Increase'] = matrix[classes].sum(axis=0)
        matrix.loc['Decrease'] = matrix['Decrease'].tolist()
        matrix.loc['Net Change'] = matrix.loc['Increase'] - matrix.loc['Decrease']
        matrix.loc['Increase', 'Decrease'] = sum(matrix['Decrease'][0:len(classes)])

        # store results
        matrices.append(matrix.copy())

    if len(matrices) > 0:
        # write matrices
        sheets = []
        mode = 'a'
        if not os.path.isfile(qaqc_path):
            mode = 'w'
        with pd.ExcelWriter(qaqc_path, mode=mode, engine='openpyxl') as writer:
            for i in range(len(matrices)):
                sheet = f"{years[i]}-{years[i+1]}-{version}" 
                matrices[i].to_excel(writer, sheet_name=sheet, index=True)
                sheets.append(sheet)

        # highlight cells that need double checking
        addStyle(qaqc_path, sheets, classes, colors)

    # merge all totals dfs
    total_df = None
    for tot_df in totals:
        if total_df is None:
            total_df = tot_df.copy()
        else:
            total_df = (
                total_df
                .merge(tot_df, on='LandCover', how='outer')
                .fillna(0.0)
            )

    # return totals
    return total_df

def totals_helper(lcc_df:pd.DataFrame, early:str, late:str, version:str) -> pd.DataFrame:
    """
    totals_helper 
    
    Calculate the total Land Cover, in acres, for the two static time periods represented in the change raster.

    Parameters
    ----------
    lcc_df : pd.DataFrame
        Table of acres of LC change classes.
    early : str
        Early year (i.e. 2014).
    late : str
        Late year (i.e. 2018).
    version : str
        Version of data being summarized (i.e. 2024ed)

    Returns
    -------
    pd.DataFrame
        Acres of total static land cover by class for the early and late dates represented in the LC change raster.
    """

    # make local copy
    df = lcc_df.copy()

    # copy no change values to late date
    df.loc[df[late].isna(), late] = df[early]

    # calculate totals from the early date
    early_tot = (
        df
        .filter(items=[early,'Acres'])
        .groupby([early])
        .sum()
        .reset_index()
        .rename(columns={
            'Acres' : f"{early}_{early[2:]}{late[2:]}_{version}",
            early : "LandCover"
        })
    )

    # calculate totals for the late date
    late_tot = (
        df
        .filter(items=[late,'Acres'])
        .groupby([late])
        .sum()
        .reset_index()
        .rename(columns={
            'Acres' : f"{late}_{early[2:]}{late[2:]}_{version}",
            late : "LandCover"
        })
    )

    # merge data
    totals = (
        early_tot
        .merge(late_tot, on='LandCover', how='outer')
        .fillna(0.0)
    )

    # return data
    return totals

def write_static_totals(df22:pd.DataFrame, df24:pd.DataFrame, qaqc_path:str):
    """
    write_static_totals _summary_

    Parameters
    ----------
    df22 : pd.DataFrame
        _description_
    df24 : pd.DataFrame
        _description_
    qaqc_path : str
        _description_
    """
    # merge LC totals
    all_totals = (
        df24
        .merge(df22, on='LandCover', how='outer')
        .fillna(0.0)
    )

    # total mapped area
    all_totals = (
        all_totals
        .set_index('LandCover')
    )

    # add totals row
    all_totals.loc['Total Acres'] = all_totals.sum(axis=0)

    # organize columns
    columns = list(all_totals)
    columns.sort()
    all_totals = all_totals[columns]

    # write totals
    with pd.ExcelWriter(qaqc_path, mode='a', engine='openpyxl') as writer:
        sheet = f"LC_Totals" 
        all_totals.to_excel(writer, sheet_name=sheet, index=True)

    # add style
    addStyle(qaqc_path, 
             sheets=[sheet], 
             classes=[], 
             sheet_type='totals')


def difference_matrices(
                qaqc_path:str,
                year1:int,
                year2:int,
                versions:list):
    """
    difference_matrices 
        For the change periods defined by the years list, difference the same change period matrices from the different versions of data.

    Parameters
    ----------
    qaqc_path : str
        Path to write results (XLSX file)
    year1 : int
        Start year of the change period.
    year2 : int
        End year of the change period.
    versions : list
        versions of data to difference. Calculations will subtract the second in the list from the first in the list (i.e. versions[0] - versions[1]).
    """
    if len(versions) != 2:
        raise Exception(f"Expected 2 versions to compare. Got {versions}")

    # read in first matrix
    df1 = (
        pd.read_excel(qaqc_path, sheet_name=f"{year1}-{year2}-{versions[0]}")
        .set_index(str(year1))
    )

    # read in second matrix
    df2 = (
        pd.read_excel(qaqc_path, sheet_name=f"{year1}-{year2}-{versions[1]}")
        .set_index(str(year1))
    )

    # store total change in most recent version
    tot_change = df1.loc['Increase', 'Decrease']

    # difference data
    df1 = df1 - df2

    # normalize by total change in most recent version
    df1 = df1 / tot_change

    # write results
    sheet = f"{year1}-{year2}_{versions[0]}-{versions[1]}" 
    with pd.ExcelWriter(qaqc_path, mode='a', engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name=sheet, index=True)

    # add style
    addStyle(qaqc_path, 
             sheets=[sheet], 
             classes=list(df1.columns[0:-1]), 
             sheet_type='difference')